import openpyxl


wb = openpyxl.load_workbook(filename='/home/pedro/Projetos/excel-python/python_excel/teste.xlsx')

print(wb.sheetnames)

for d in wb['Sheet1'].iter_rows(values_only = True):
    nome = str(d[0]).title()
    
    if d[1] != None and len(str(d[1])) <= 9:
        rg = {'RG': d[1]}
        cpf = None
    else:
        cpf = {'cpf': d[1]}
        rg = None

    bairro = {'bairro': d[2]}
    rua = {'rua': d[3]}
    numero = {'numero': d[4]}
    
    if d[5] is None:
        complemento = {'complemento': '-'}
    else:
        complemento = {'complemento': d[5]}

    cidade = {'cidade': 'Limeira'}
    estado = {'estado': 'São Paulo'}

    print(nome, rg, cpf, bairro, rua, numero, complemento, cidade, estado)

    # Para Django
'''    Pessoa(
        nome=nome,
        cpf=cpf,
        rg=rg,
        logradouro=logradouro,
        numero=numero,
        bairro=bairro,
        cidade='Limeira',
        estado='SP',
        atualizado_por=User(1),
        criado_por=User(1)
    ).save()'''
